import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import xlrd
import pandas as pd

# getting all file names from student list directory 
students_path = "./student-lists"
dir_list = ["KET 1 - K41.xls"]
#dir_list = os.listdir(students_path) 

for list in dir_list:
    # get student data as a data frame
    student_data = pd.read_html(f'{students_path}/{list}')[0]
    # reading from templates
    template_path = './templates/extra-classes/TUTORING COURSE OUTLINE - KET 1.xlsx'
    output_path = './templates/'
    wb_obj = openpyxl.load_workbook(template_path)
    # generate new file names based on student lists
    file_name = f'{output_path}{list.split(".")[0]}_extra_class.xlsx'
    # get active sheet from template
    sheet_obj = wb_obj.active 
    # modify title into appropriate class
    name = sheet_obj.cell(row = 1, column = 1) 
    name.value = f'{name.value}  {list.split("-")[1].split(".")[0].upper()}'
    #print(name.value)
    # modify student columns
    # col pointer starts at 3 (hard coded atm)
    col_pointer = 4
    # setting student detail font to bold
    student_font = Font(name="Time News Roman", size=12, bold=True)
    for i, student in enumerate(student_data['Tên']):
        # writing student index
        index_cell = sheet_obj.cell(2,col_pointer)
        index_cell.value = i + 1
        index_cell.font = student_font
        # writing students' names
        if type(sheet_obj.cell(3, col_pointer)).__name__ == 'MergedCell':
            student_cell = sheet_obj.unmerge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        student_cell = sheet_obj.cell(3,col_pointer)
        student_cell.value = student
        student_cell.font = student_font
        student_cell = sheet_obj.merge_cells(start_row=3,start_column=col_pointer, end_row=3, end_column=col_pointer+2)
        # writing skill cells
        listening_cell = sheet_obj.cell(4,col_pointer)
        listening_cell.value = "List."
        listening_cell.font = student_font
        reading_writing_cell = sheet_obj.cell(4,col_pointer+1)
        reading_writing_cell.value = "R&W"
        reading_writing_cell.font = student_font
        vocab_cell = sheet_obj.cell(4,col_pointer+2)
        vocab_cell.value = "Vocab"
        vocab_cell.font = student_font
        col_pointer = col_pointer+3
    # page setup
    sheet_obj.page_setup.orientation = "landscape"
    sheet_obj.page_setup.scale = 75
    sheet_obj.sheet_properties.pageSetUpPr.fitToPage = False
    sheet_obj.print_options.horizontalCentered = False
    sheet_obj.print_options.verticalCentered  = False
    sheet_obj.page_margins.top = 0.314
    sheet_obj.page_margins.left = 0.17
    sheet_obj.page_margins.bottom = 0.2
    sheet_obj.page_margins.right = 0.16
    sheet_obj._print_area = None
    wb_obj.save(file_name)
